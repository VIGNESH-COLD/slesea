import os
import uuid
from flask import Flask, render_template, request, redirect, url_for, send_from_directory, session, flash
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'supersecretkey'

# ---------------------------- Admin Login ----------------------------
@app.route('/admin-login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username == 'Admin@suvai_organics' and password == 'suvai_1209':
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid username or password', 'error')

    return render_template('admin-login.html')

@app.route('/admin-logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/admin-dashboard')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    return render_template('admin-dashboard.html')

# ---------------------------- Config ----------------------------
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['GALLERY_FOLDER'] = 'static/uploads/gallery'
app.config['EXCEL_FILE'] = 'data/products.xlsx'
app.config['GALLERY_FILE'] = 'data/gallery.xlsx'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024  # 2MB

# Create required directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['GALLERY_FOLDER'], exist_ok=True)
os.makedirs('data', exist_ok=True)

# Initialize Excel files if not exists
if not os.path.exists(app.config['EXCEL_FILE']):
    wb = Workbook()
    ws = wb.active
    ws.append(['sku', 'name', 'description', 'price', 'category', 'stock', 'image_filename', 'created_at'])
    wb.save(app.config['EXCEL_FILE'])

if not os.path.exists(app.config['GALLERY_FILE']):
    wb = Workbook()
    ws = wb.active
    ws.append(['id', 'filename', 'caption', 'category', 'created_at'])
    wb.save(app.config['GALLERY_FILE'])

# ---------------------------- Helpers ----------------------------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def get_products():
    if not os.path.exists(app.config['EXCEL_FILE']):
        return []
    
    wb = load_workbook(app.config['EXCEL_FILE'])
    ws = wb.active
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Check if SKU exists (not empty row)
            products.append({
                'sku': row[0],
                'name': row[1],
                'description': row[2],
                'price': row[3],
                'category': row[4],
                'stock': row[5],
                'image_filename': row[6],
                'created_at': row[7]
            })
    return products

def save_product_to_excel(product_data):
    wb = load_workbook(app.config['EXCEL_FILE'])
    ws = wb.active
    ws.append([
        product_data['sku'],
        product_data['name'],
        product_data['description'],
        product_data['price'],
        product_data['category'],
        product_data['stock'],
        product_data['image_filename'],
        product_data['created_at']
    ])
    wb.save(app.config['EXCEL_FILE'])

def get_gallery_items():
    if not os.path.exists(app.config['GALLERY_FILE']):
        return []
    
    wb = load_workbook(app.config['GALLERY_FILE'])
    ws = wb.active
    gallery = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Check if ID exists (not empty row)
            gallery.append({
                'id': row[0],
                'filename': row[1],
                'caption': row[2],
                'category': row[3],
                'created_at': row[4]
            })
    return gallery

def save_gallery_item(item_data):
    wb = load_workbook(app.config['GALLERY_FILE'])
    ws = wb.active
    ws.append([
        item_data['id'],
        item_data['filename'],
        item_data['caption'],
        item_data['category'],
        item_data['created_at']
    ])
    wb.save(app.config['GALLERY_FILE'])

def delete_gallery_item(item_id):
    wb = load_workbook(app.config['GALLERY_FILE'])
    ws = wb.active
    
    # Find the row with the matching ID
    row_index = None
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == item_id:
            row_index = idx
            break
    
    if row_index:
        ws.delete_rows(row_index)
        wb.save(app.config['GALLERY_FILE'])
        return True
    return False

# ---------------------------- Public Routes ----------------------------
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/contact')
def contact():
    return render_template('contact.html')

@app.route('/gallery')
def gallery():
    gallery_items = get_gallery_items()
    # Extract unique categories
    categories = sorted(set(item['category'] for item in gallery_items if item['category']))
    return render_template('gallery.html', gallery_items=gallery_items, categories=categories)

@app.route('/natural-farming')
def natural_farming():
    return render_template('natural-farming.html')

@app.route('/farmers')
def farmers():
    return render_template('farmers.html')

@app.route('/partners')
def partners():
    return render_template('partners.html')

@app.route('/register')
def register_main():
    return render_template('registration.html')

@app.route('/register-customer')
def register_customer():
    return render_template('register-customer.html')

@app.route('/register-partner')
def register_partner():
    return render_template('register-partner.html')

@app.route('/register-farmer')
def register_farmer():
    return render_template('register-farmer.html')

@app.route('/products')
def products():
    products = get_products()
    categories = sorted(set(p['category'] for p in products if p['category']))
    return render_template('products.html', products=products, categories=categories)

# ---------------------------- Admin Routes ----------------------------
@app.route('/products/upload', methods=['GET', 'POST'])
def products_upload():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    if request.method == 'POST':
        # Get form data
        sku = request.form.get('sku')
        name = request.form.get('name')
        description = request.form.get('description')
        price = request.form.get('price')
        category = request.form.get('category')
        stock = request.form.get('stock')
        image_file = request.files.get('image')
        
        # Validate required fields
        if not sku or not name or not price:
            flash('SKU, Name, and Price are required fields', 'error')
            return redirect(request.url)
        
        try:
            price = float(price)
        except ValueError:
            flash('Invalid price format', 'error')
            return redirect(request.url)
        
        # Handle image upload
        image_filename = None
        if image_file and image_file.filename != '':
            if not allowed_file(image_file.filename):
                flash('Invalid file type. Allowed: PNG, JPG, JPEG, GIF', 'error')
                return redirect(request.url)
            
            filename = secure_filename(image_file.filename)
            unique_filename = f"{uuid.uuid4().hex}_{filename}"
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            image_file.save(image_path)
            image_filename = unique_filename
        
        # Create product data
        product_data = {
            'sku': sku,
            'name': name,
            'description': description,
            'price': price,
            'category': category,
            'stock': stock,
            'image_filename': image_filename,
            'created_at': datetime.now().isoformat()
        }
        
        # Save to Excel
        try:
            save_product_to_excel(product_data)
            flash('Product uploaded successfully!', 'success')
            return redirect(url_for('products'))
        except Exception as e:
            flash(f'Error saving product: {str(e)}', 'error')
    
    return render_template('products-upload.html')

@app.route('/gallery-manage', methods=['GET', 'POST'])
def gallery_manage():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    if request.method == 'POST':
        # Handle image upload
        image_file = request.files.get('image')
        caption = request.form.get('caption', '')
        category = request.form.get('category', 'Uncategorized')
        
        if not image_file or image_file.filename == '':
            flash('No image selected for upload', 'error')
            return redirect(request.url)
        
        if not allowed_file(image_file.filename):
            flash('Invalid file type. Allowed: PNG, JPG, JPEG, GIF', 'error')
            return redirect(request.url)
        
        # Save image
        filename = secure_filename(image_file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        image_path = os.path.join(app.config['GALLERY_FOLDER'], unique_filename)
        image_file.save(image_path)
        
        # Create gallery item data
        item_data = {
            'id': str(uuid.uuid4()),
            'filename': unique_filename,
            'caption': caption,
            'category': category,
            'created_at': datetime.now().isoformat()
        }
        
        # Save to gallery Excel
        try:
            save_gallery_item(item_data)
            flash('Gallery image uploaded successfully!', 'success')
            return redirect(url_for('gallery_manage'))
        except Exception as e:
            flash(f'Error saving gallery item: {str(e)}', 'error')
    
    # For GET requests, show gallery items
    gallery_items = get_gallery_items()
    return render_template('gallery-manage.html', gallery_items=gallery_items)

@app.route('/gallery-manage/delete/<item_id>', methods=['POST'])
def delete_gallery_item_route(item_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    if delete_gallery_item(item_id):
        flash('Gallery item deleted successfully!', 'success')
    else:
        flash('Gallery item not found', 'error')
    
    return redirect(url_for('gallery_manage'))

@app.route('/farmers-manage')
def farmers_manage():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    return render_template('farmers-manage.html')

@app.route('/partners-manage')
def partners_manage():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    return render_template('partners-manage.html')

# ---------------------------- Safe Static File Serving ----------------------------
@app.route('/static/<path:filename>')
def custom_static(filename):
    return send_from_directory('static', filename)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')